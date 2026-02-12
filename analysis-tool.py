"""
Mountain Power — Analysis File Generator (Web App)
Streamlit app for transforming AI-generated listing files into full analysis files.
Deploy free on Streamlit Community Cloud.
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Analysis File Generator — Mountain Power",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

ANALYSIS_COLUMNS = [
    "ITEM", "Listing SKU", "Blocked", "Blocked Notes", "Brand", "Product Type",
    "Deposco Category", "Amazon Title", "eBay Title", "Amazon Trademark Brand List",
    "Amazon Category", "Store Category", "Store Featured Category", "Keywords",
    "Amazon Description", "Applications", "Replaces", "Fitment", "Store Fitment",
    "Division Listing Type", "Length", "Width", "Height", "Weight", "Images",
    "Unit Cost", "BIN", "Vendor", "Source", "Partslink Numbers",
    "Manufacturer Part Number", "Interchange Part Number", "Other Part Number",
    "OEM Interchange Part Number 1", "OEM Interchange Part Number 2",
    "OEM Interchange Part Number 3", "OEM Interchange Part Number 4",
    "OEM Interchange Part Number 5", "OEM Interchange Part Number 6",
    "OEM Interchange Part Number 7", "OEM Interchange Part Number 8",
    "OEM Interchange Part Number 9"
]

SUPPLEMENTAL_FIELDS = [
    ("Brand", "text", "e.g. Rareelectrical"),
    ("Product Type", "text", "e.g. Turbocharger"),
    ("Deposco Category", "text", "e.g. Turbocharger"),
    ("Amazon Trademark Brand List", "text", "e.g. CUMMINS,MAXIFORCE"),
    ("Amazon Category", "text", "e.g. Automotive"),
    ("Store Category", "text", "e.g. Turbochargers"),
    ("Store Featured Category", "text", ""),
    ("Division Listing Type", "select", ["", "Primary", "Secondary"]),
    ("Length", "number", "inches"),
    ("Width", "number", "inches"),
    ("Height", "number", "inches"),
    ("Weight", "number", "lbs"),
    ("Images", "text", "Image URLs"),
    ("Unit Cost", "number", "$ amount"),
    ("BIN", "number", "Retail Price $"),
    ("Vendor", "text", "e.g. Maxiforce"),
    ("Source", "text", "e.g. Aftermarket"),
    ("Partslink Numbers", "text", ""),
]

ITEM_SHEET_DEFAULTS = {
    "Drop Ship": 1, "Reorder Lead Time": 0, "Minimum Order Quantity": 0,
    "Reorder Point": 0, "Reorder Quantity": 0, "Inventory Tracking Enabled": True,
    "Shippable Flag": True, "FL Reorder Point": 0, "FL Reorder Quantity": 0,
    "Default Receive Quantity": 1
}


# ═══════════════════════════════════════════════════════════════════════════════
# CUSTOM CSS
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    /* Global */
    .stApp { background-color: #0a0a0a; }
    section[data-testid="stSidebar"] { background-color: #111; }
    .block-container { padding-top: 2rem; max-width: 1100px; }

    /* Header bar */
    .app-header {
        background: linear-gradient(135deg, #1a1a1a 0%, #111 100%);
        border: 1px solid #222;
        border-radius: 16px;
        padding: 20px 28px;
        margin-bottom: 24px;
        display: flex;
        align-items: center;
        justify-content: space-between;
    }
    .app-title {
        font-family: 'DM Sans', sans-serif;
        font-size: 22px;
        font-weight: 700;
        color: #f5f5f5;
        margin: 0;
    }
    .app-subtitle {
        font-family: 'DM Sans', sans-serif;
        font-size: 13px;
        color: #666;
        margin: 0;
    }

    /* Step indicator */
    .step-bar {
        display: flex;
        gap: 6px;
        margin-bottom: 20px;
    }
    .step-pill {
        padding: 6px 16px;
        border-radius: 20px;
        font-family: 'DM Sans', sans-serif;
        font-size: 12px;
        font-weight: 600;
    }
    .step-active { background: #b4530920; color: #f59e0b; border: 1px solid #f59e0b40; }
    .step-done { background: #16653420; color: #22c55e; border: 1px solid #22c55e30; }
    .step-pending { background: #1a1a1a; color: #555; border: 1px solid #222; }

    /* Cards */
    .metric-card {
        background: #141414;
        border: 1px solid #222;
        border-radius: 12px;
        padding: 16px;
        text-align: center;
    }
    .metric-value {
        font-family: 'DM Sans', sans-serif;
        font-size: 28px;
        font-weight: 700;
        color: #f59e0b;
    }
    .metric-label {
        font-family: 'DM Sans', sans-serif;
        font-size: 11px;
        color: #666;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    /* Status badges */
    .badge-green { background: #16653420; color: #22c55e; border: 1px solid #22c55e30; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; font-family: 'DM Sans'; display: inline-block; }
    .badge-amber { background: #b4530920; color: #f59e0b; border: 1px solid #f59e0b30; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; font-family: 'DM Sans'; display: inline-block; }
    .badge-red { background: #7f1d1d30; color: #ef4444; border: 1px solid #ef444430; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; font-family: 'DM Sans'; display: inline-block; }

    /* Field list panels */
    .field-panel {
        border-radius: 10px;
        padding: 14px;
        font-family: 'DM Sans', sans-serif;
        font-size: 12px;
    }
    .field-panel-green { background: #0d1f0d; border: 1px solid #16653430; }
    .field-panel-amber { background: #1f1a0d; border: 1px solid #b4530930; }
    .field-panel-red { background: #1f0d0d; border: 1px solid #7f1d1d40; }
    .field-panel h4 { font-size: 10px; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }
    .field-item { padding: 3px 0; color: #999; display: flex; justify-content: space-between; }
    .field-pct-green { color: #22c55e; }
    .field-pct-amber { color: #f59e0b; }

    /* Success banner */
    .success-banner {
        text-align: center;
        padding: 30px;
        background: linear-gradient(135deg, #0d1f0d 0%, #111 100%);
        border: 1px solid #16653430;
        border-radius: 16px;
        margin-bottom: 20px;
    }
    .success-icon { font-size: 48px; margin-bottom: 8px; }
    .success-title { font-family: 'DM Sans'; font-size: 24px; font-weight: 700; color: #f5f5f5; margin: 0; }
    .success-sub { font-family: 'DM Sans'; font-size: 13px; color: #666; margin: 4px 0 0 0; }

    /* Section headers */
    .section-label {
        font-family: 'DM Sans', sans-serif;
        font-size: 10px;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 1.5px;
        color: #555;
        margin-bottom: 10px;
    }

    /* Hide streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    /* Style file uploader */
    [data-testid="stFileUploader"] {
        border: 2px dashed #2a2a2a;
        border-radius: 12px;
        padding: 8px;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: #444;
    }

    /* Style dataframes */
    [data-testid="stDataFrame"] {
        border: 1px solid #222;
        border-radius: 10px;
    }

    /* Buttons */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #f59e0b, #d97706) !important;
        color: #000 !important;
        font-weight: 700 !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 10px 32px !important;
        font-size: 15px !important;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #fbbf24, #f59e0b) !important;
    }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# DATA PROCESSING (same logic as desktop version)
# ═══════════════════════════════════════════════════════════════════════════════

@st.cache_data
def read_uploaded_file(uploaded_file):
    """Read uploaded file into dict of sheet_name -> list of dicts."""
    try:
        xl = pd.ExcelFile(uploaded_file)
        sheets = {}
        for name in xl.sheet_names:
            df = xl.parse(name)
            df = df.where(pd.notnull(df), None)
            sheets[name] = df.to_dict("records")
        return {"sheets": sheets, "sheet_names": xl.sheet_names, "name": uploaded_file.name}
    except Exception:
        return None


def detect_best_sheet(file_data):
    names = file_data["sheet_names"]
    if len(names) == 1:
        return names[0]
    priority = ["rithum upload", "listings", "listing", "data"]
    for p in priority:
        for n in names:
            if p in n.lower():
                return n
    best, max_cols = names[0], 0
    for n in names:
        data = file_data["sheets"].get(n, [])
        if data and isinstance(data[0], dict):
            cols = len(data[0].keys())
            if cols > max_cols:
                max_cols = cols
                best = n
    return best


def detect_fields(data):
    if not data:
        return {"present": [], "missing": list(ANALYSIS_COLUMNS), "partial": []}
    cols = set()
    for row in data:
        cols.update(row.keys())
    col_lower_map = {c.lower().strip(): c for c in cols}
    present, missing, partial = [], [], []
    for col in ANALYSIS_COLUMNS:
        matched = col_lower_map.get(col.lower().strip())
        if matched:
            filled = sum(1 for r in data if r.get(matched) is not None and str(r.get(matched, "")).strip())
            pct = filled / len(data) if data else 0
            if pct > 0.8:
                present.append((col, matched, pct))
            elif pct > 0:
                partial.append((col, matched, pct))
            else:
                missing.append(col)
        else:
            missing.append(col)
    return {"present": present, "missing": missing, "partial": partial}


def get_unique_items(data):
    items = set()
    for r in data:
        item = r.get("ITEM") or r.get("Item") or r.get("item")
        if item:
            items.add(str(item).strip())
    return sorted(items)


def build_analysis_rows(source_data, global_defaults, supp_data=None):
    supp_map = {}
    if supp_data:
        for r in supp_data:
            key = r.get("ITEM") or r.get("Item") or r.get("SKU") or r.get("Listing SKU")
            if not key:
                vals = list(r.values())
                key = vals[0] if vals else None
            if key:
                supp_map[str(key).strip().upper()] = r
    rows = []
    for row in source_data:
        out = {}
        row_lower = {k.lower().strip(): k for k in row.keys()}
        for col in ANALYSIS_COLUMNS:
            src_key = row_lower.get(col.lower().strip())
            val = None
            if src_key and row.get(src_key) is not None and str(row.get(src_key, "")).strip():
                val = row[src_key]
            if val is None:
                item_key = str(row.get("ITEM") or row.get("Item") or "").strip().upper()
                supp = supp_map.get(item_key)
                if supp:
                    supp_lower = {k.lower().strip(): k for k in supp.keys()}
                    supp_key = supp_lower.get(col.lower().strip())
                    if supp_key and supp.get(supp_key) is not None and str(supp.get(supp_key, "")).strip():
                        val = supp[supp_key]
            if val is None and col in global_defaults and str(global_defaults[col]).strip():
                val = global_defaults[col]
            out[col] = val if val is not None else ""
        if not out.get("Blocked"):
            out["Blocked"] = False
        rows.append(out)
    return rows


def build_item_vendor_sheet(rows):
    seen, out = set(), []
    for r in rows:
        item = r.get("ITEM")
        if item and item not in seen:
            seen.add(item)
            out.append({"Fulfillment Type": "", "Item": item, "Trading Partner": r.get("Vendor", ""),
                        "SKU/UPC": item, "Unit Cost": r.get("Unit Cost", ""), "Is Preferred Vendor": True, "Quantity": ""})
    return out


def build_item_sheet(rows, harmonized_code, origin_country):
    seen, out = set(), []
    for r in rows:
        item = r.get("ITEM")
        if item and item not in seen:
            seen.add(item)
            out.append({"ID": "", "Number": item, "Name": item,
                        "Long Description": r.get("eBay Title", ""), "Unit Cost": r.get("Unit Cost", ""),
                        **ITEM_SHEET_DEFAULTS, "Trading Partner": r.get("Vendor", ""),
                        "Retail Price": r.get("BIN", ""),
                        "Product Category": r.get("Product Type") or r.get("Deposco Category", ""),
                        "Harmonized Code": harmonized_code, "Short Description": "", "Origin Country": origin_country})
    return out


def build_pack_sheet(rows):
    seen, out = set(), []
    for r in rows:
        item = r.get("ITEM")
        if item and item not in seen:
            seen.add(item)
            out.append({"Pack Key": f"{item}--Each--1", "Item": item, "Pack Type": "Each", "Quantity": 1,
                        "Length": r.get("Length", ""), "Length Uom": "Inch",
                        "Width": r.get("Width", ""), "Width Uom": "Inch",
                        "Height": r.get("Height", ""), "Height Uom": "Inch",
                        "Volume": "", "Volume Uom": "",
                        "Weight": r.get("Weight", ""), "Weight Uom": "Pound"})
    return out


def build_item_upc_sheet(rows):
    return [{"ITEM": r.get("ITEM", ""), "UPC": r.get("Listing SKU", ""), "Source": "Listings"} for r in rows]


def export_to_bytes(analysis_rows, harmonized_code, origin_country):
    """Create the full multi-sheet Excel file and return as bytes."""
    wb = openpyxl.Workbook()
    header_font = Font(name="Aptos Narrow", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2d2d2d")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Aptos Narrow", size=10)
    thin_border = Border(
        left=Side(style="thin", color="3a3a3a"), right=Side(style="thin", color="3a3a3a"),
        top=Side(style="thin", color="3a3a3a"), bottom=Side(style="thin", color="3a3a3a")
    )

    def write_sheet(ws, data, columns=None):
        if not data:
            return
        if columns is None:
            columns = list(data[0].keys())
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for ri, row in enumerate(data, 2):
            for ci, col in enumerate(columns, 1):
                val = row.get(col, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        for ci, col in enumerate(columns, 1):
            max_len = len(str(col))
            for ri in range(2, min(len(data) + 2, 20)):
                val = str(data[ri - 2].get(col, ""))[:50]
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 45)
        ws.freeze_panes = "A2"
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data) + 1}"

    ws1 = wb.active
    ws1.title = "Rithum Upload"
    write_sheet(ws1, analysis_rows, ANALYSIS_COLUMNS)

    iv_data = build_item_vendor_sheet(analysis_rows)
    if iv_data:
        write_sheet(wb.create_sheet("ItemVendor"), iv_data)

    item_data = build_item_sheet(analysis_rows, harmonized_code, origin_country)
    if item_data:
        write_sheet(wb.create_sheet("Item"), item_data)

    pack_data = build_pack_sheet(analysis_rows)
    if pack_data:
        write_sheet(wb.create_sheet("Pack"), pack_data)

    upc_data = build_item_upc_sheet(analysis_rows)
    if upc_data:
        write_sheet(wb.create_sheet("ItemUPC"), upc_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def render_step_bar(current):
    steps = ["Upload", "Configure", "Deposco", "Export"]
    pills = []
    for i, name in enumerate(steps):
        if i == current:
            pills.append(f'<span class="step-pill step-active">● {i+1}. {name}</span>')
        elif i < current:
            pills.append(f'<span class="step-pill step-done">✓ {name}</span>')
        else:
            pills.append(f'<span class="step-pill step-pending">{i+1}. {name}</span>')
    st.markdown(f'<div class="step-bar">{"".join(pills)}</div>', unsafe_allow_html=True)


def render_field_scan(analysis):
    present = analysis["present"]
    partial = analysis["partial"]
    missing = analysis["missing"]

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<span class="badge-green">● {len(present)} Complete</span>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<span class="badge-amber">● {len(partial)} Partial</span>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<span class="badge-red">● {len(missing)} Missing</span>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        if present:
            items_html = "".join(
                f'<div class="field-item"><span>{name}</span><span class="field-pct-green">{int(pct*100)}%</span></div>'
                for name, _, pct in present
            )
            st.markdown(f'<div class="field-panel field-panel-green"><h4 style="color:#22c55e">Complete</h4>{items_html}</div>', unsafe_allow_html=True)
    with c2:
        if partial:
            items_html = "".join(
                f'<div class="field-item"><span>{name}</span><span class="field-pct-amber">{int(pct*100)}%</span></div>'
                for name, _, pct in partial
            )
            st.markdown(f'<div class="field-panel field-panel-amber"><h4 style="color:#f59e0b">Partial</h4>{items_html}</div>', unsafe_allow_html=True)
    with c3:
        if missing:
            items_html = "".join(f'<div class="field-item"><span>{name}</span></div>' for name in missing)
            st.markdown(f'<div class="field-panel field-panel-red"><h4 style="color:#ef4444">Missing</h4>{items_html}</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ═══════════════════════════════════════════════════════════════════════════════

if "step" not in st.session_state:
    st.session_state.step = 0
if "source_data" not in st.session_state:
    st.session_state.source_data = None
if "analysis" not in st.session_state:
    st.session_state.analysis = None
if "source_name" not in st.session_state:
    st.session_state.source_name = ""
if "output_rows" not in st.session_state:
    st.session_state.output_rows = None
if "defaults" not in st.session_state:
    st.session_state.defaults = {}


# ═══════════════════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="app-header">
    <div>
        <p class="app-title">⚡ Analysis File Generator</p>
        <p class="app-subtitle">Mountain Power — Listing Pipeline</p>
    </div>
    <div style="font-family: 'JetBrains Mono', monospace; font-size: 11px; color: #555;">
        42 fields • 5 sheets
    </div>
</div>
""", unsafe_allow_html=True)

render_step_bar(st.session_state.step)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 0: UPLOAD
# ═══════════════════════════════════════════════════════════════════════════════

if st.session_state.step == 0:
    st.markdown("### Upload Source File")
    st.markdown('<p style="color:#888; font-size:14px;">Upload the AI-generated listing file. The tool will scan all 42 analysis columns and show what\'s present, partial, or missing.</p>', unsafe_allow_html=True)

    uploaded = st.file_uploader("Drop your source listing file", type=["xlsx", "xls", "csv"], key="source_upload")

    if uploaded:
        file_data = read_uploaded_file(uploaded)
        if file_data:
            st.session_state.source_name = file_data["name"]

            # Sheet selector
            best_sheet = detect_best_sheet(file_data)
            if len(file_data["sheet_names"]) > 1:
                selected_sheet = st.selectbox(
                    "Select the sheet containing listing data:",
                    file_data["sheet_names"],
                    index=file_data["sheet_names"].index(best_sheet),
                    key="sheet_select"
                )
            else:
                selected_sheet = best_sheet

            data = file_data["sheets"][selected_sheet]
            st.session_state.source_data = data
            analysis = detect_fields(data)
            st.session_state.analysis = analysis

            items = get_unique_items(data)
            st.markdown(f'<p style="color:#888; font-family: JetBrains Mono, monospace; font-size:12px;">{len(data)} rows • {len(items)} unique items • Sheet: {selected_sheet}</p>', unsafe_allow_html=True)

            # Scan results
            st.markdown('<div class="section-label">Field Scan Results</div>', unsafe_allow_html=True)
            render_field_scan(analysis)

            # Preview
            st.markdown("---")
            st.markdown('<div class="section-label">Source Data Preview</div>', unsafe_allow_html=True)
            preview_cols = [c for c in ["ITEM", "Listing SKU", "Brand", "Amazon Title", "Unit Cost", "Vendor", "Weight"] if c in data[0]] if data else []
            if preview_cols:
                preview_df = pd.DataFrame(data[:8])[preview_cols]
                st.dataframe(preview_df, use_container_width=True, hide_index=True)

            st.markdown("")
            if st.button("Continue to Configuration →", type="primary", use_container_width=False):
                st.session_state.step = 1
                st.rerun()
        else:
            st.error("Could not read file. Make sure it's a valid Excel file.")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1: CONFIGURE
# ═══════════════════════════════════════════════════════════════════════════════

elif st.session_state.step == 1:
    st.markdown("### Configure Missing Fields")
    st.markdown('<p style="color:#888; font-size:14px;">Fill in global defaults for missing fields, or attach a supplemental file with per-item values (matched by ITEM column).</p>', unsafe_allow_html=True)

    # Supplemental file
    st.markdown('<div class="section-label">Optional: Supplemental Data File</div>', unsafe_allow_html=True)
    st.markdown('<p style="color:#666; font-size:12px;">Attach a file with per-item prices, dimensions, etc. Rows will be matched by ITEM / SKU column.</p>', unsafe_allow_html=True)
    supp_upload = st.file_uploader("Drop supplemental data file (optional)", type=["xlsx", "xls", "csv"], key="supp_upload")

    supp_data = None
    if supp_upload:
        supp_file = read_uploaded_file(supp_upload)
        if supp_file:
            supp_sheet = supp_file["sheet_names"][0]
            supp_data = supp_file["sheets"][supp_sheet]
            st.success(f"✓ Loaded {len(supp_data)} rows from \"{supp_sheet}\"")
    st.session_state["supp_data"] = supp_data

    # Defaults
    st.markdown("---")
    analysis = st.session_state.analysis
    missing_and_partial = set(analysis["missing"]) | set(p[0] for p in analysis["partial"])
    relevant = [f for f in SUPPLEMENTAL_FIELDS if f[0] in missing_and_partial]

    if not relevant:
        st.success("✓ All supplemental fields are already populated in the source file.")
    else:
        st.markdown('<div class="section-label">Fill Missing / Partial Fields (Applied Globally)</div>', unsafe_allow_html=True)
        cols = st.columns(3)
        defaults = {}
        for i, (key, ftype, placeholder) in enumerate(relevant):
            with cols[i % 3]:
                if ftype == "select":
                    val = st.selectbox(key, placeholder, key=f"def_{key}")
                elif ftype == "number":
                    val = st.text_input(key, placeholder=placeholder, key=f"def_{key}")
                else:
                    val = st.text_input(key, placeholder=placeholder, key=f"def_{key}")
                if val:
                    defaults[key] = val
        st.session_state.defaults = defaults

    st.markdown("")
    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button("← Back"):
            st.session_state.step = 0
            st.rerun()
    with c2:
        if st.button("Continue to Deposco Settings →", type="primary"):
            # Collect defaults from widgets
            for key, ftype, _ in relevant:
                val = st.session_state.get(f"def_{key}", "")
                if val:
                    st.session_state.defaults[key] = val
            st.session_state.step = 2
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2: DEPOSCO
# ═══════════════════════════════════════════════════════════════════════════════

elif st.session_state.step == 2:
    st.markdown("### Deposco & Item Settings")
    st.markdown('<p style="color:#888; font-size:14px;">Configure settings for the Item, ItemVendor, Pack, and ItemUPC sheets that will be auto-generated.</p>', unsafe_allow_html=True)

    st.markdown('<div class="section-label">Deposco / Item Sheet Defaults</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        harmonized_code = st.text_input("Harmonized Code (HTS)", placeholder="e.g. 8414.59.6590", key="hts")
    with c2:
        origin_country = st.text_input("Country of Origin", placeholder="e.g. CN", key="origin")

    # Summary
    st.markdown("---")
    st.markdown('<div class="section-label">Auto-Generated Sheets Summary</div>', unsafe_allow_html=True)
    data = st.session_state.source_data
    items = get_unique_items(data) if data else []

    c1, c2, c3, c4, c5 = st.columns(5)
    sheets_info = [
        ("📋", "Rithum Upload", f"{len(data)} rows"),
        ("🏭", "ItemVendor", f"{len(items)} items"),
        ("📦", "Item", f"{len(items)} records"),
        ("📐", "Pack", f"{len(items)} records"),
        ("🏷️", "ItemUPC", f"{len(data)} rows"),
    ]
    for col, (icon, name, desc) in zip([c1, c2, c3, c4, c5], sheets_info):
        with col:
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size:24px">{icon}</div>
                <div style="font-family:'DM Sans'; font-size:13px; font-weight:700; color:#e4e4e7; margin:4px 0 2px;">{name}</div>
                <div style="font-family:'DM Sans'; font-size:11px; color:#666;">{desc}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("")
    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button("← Back", key="back2"):
            st.session_state.step = 1
            st.rerun()
    with c2:
        if st.button("⚡ Generate Analysis File", type="primary"):
            with st.spinner("Generating..."):
                supp_data = st.session_state.get("supp_data")
                rows = build_analysis_rows(data, st.session_state.defaults, supp_data)
                st.session_state.output_rows = rows
                st.session_state.harmonized_code = harmonized_code
                st.session_state.origin_country = origin_country
                st.session_state.step = 3
                st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3: EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

elif st.session_state.step == 3:
    rows = st.session_state.output_rows
    items = get_unique_items(rows)

    st.markdown(f"""
    <div class="success-banner">
        <div class="success-icon">✅</div>
        <p class="success-title">Analysis File Ready</p>
        <p class="success-sub">{len(rows)} listing rows  •  {len(items)} unique items  •  5 sheets</p>
    </div>
    """, unsafe_allow_html=True)

    # Preview
    st.markdown('<div class="section-label">Rithum Upload Preview</div>', unsafe_allow_html=True)
    preview_cols = ["ITEM", "Listing SKU", "Brand", "Amazon Title", "Unit Cost", "Vendor", "Weight"]
    preview_df = pd.DataFrame(rows[:8])[preview_cols]
    st.dataframe(preview_df, use_container_width=True, hide_index=True)

    # Sheet counts
    st.markdown('<div class="section-label">Sheet Summary</div>', unsafe_allow_html=True)
    c1, c2, c3, c4, c5 = st.columns(5)
    counts = [
        ("Rithum Upload", len(rows)),
        ("ItemVendor", len(items)),
        ("Item", len(items)),
        ("Pack", len(items)),
        ("ItemUPC", len(rows)),
    ]
    for col, (name, count) in zip([c1, c2, c3, c4, c5], counts):
        with col:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{count}</div>
                <div class="metric-label">{name}</div>
            </div>""", unsafe_allow_html=True)

    # Export
    st.markdown("")
    harmonized = st.session_state.get("harmonized_code", "")
    origin = st.session_state.get("origin_country", "")
    buf = export_to_bytes(rows, harmonized, origin)

    base_name = st.session_state.source_name.replace(".xlsx", "").replace(".xls", "").replace(".csv", "")
    file_name = f"{base_name}_ANALYSIS.xlsx"

    c1, c2 = st.columns([3, 1])
    with c1:
        st.download_button(
            label="💾  Download .xlsx",
            data=buf,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with c2:
        if st.button("🔄 Start Over"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("---")
st.markdown(
    '<p style="text-align:center; color:#444; font-family: JetBrains Mono, monospace; font-size:11px;">'
    'Mountain Power • Analysis File Generator v1.0 • 42 fields • 5 sheets</p>',
    unsafe_allow_html=True
)